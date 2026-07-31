"""EPA eGRID subregion factors + the WUE benchmark table — the derivation's inputs (#1082).

The committed ``reference`` factor tables the footprint derivation (#1083) multiplies the
electricity figure by, so the electricity → CO2e / source-mix / water conversion reads
authoritative published factors rather than a hand-maintained constant. Two live here; the
third (inference energy, #1643/F2) has its own module but is written by the same command and
shares this one's README:

1. **EPA eGRID** (:func:`fetch_egrid_factors`) — the annual Emissions & Generation Resource
   Integrated Database subregion file, pulled through the shared ``cached_get`` machinery.
   We download the public workbook (no API key), read the subregion sheet (``SRL<yy>``),
   and reduce each subregion to its CO2-equivalent output rate (lb/MWh) + generation mix.
   Columns are selected **by field code** (``SUBRGN``, ``SRC2ERTA``, the ``SR*PR`` mix
   columns) resolved from the sheet's own header row — never by index — so a column
   re-order between vintages doesn't silently mis-read. The reduced rows (not the ~20 MB
   xlsx) are what land in the cache/fixture, keeping tests hermetic.

2. **WUE benchmarks** (:func:`build_wue_table`) — a hand-curated Water Usage Effectiveness
   table (AWS's published WUE / EPRI / Uptime Institute / Macknick et al.), transcribed once
   with each row's dated source and band. Not a live pull; it's an in-code canonical the
   ``--write`` emits so it stays regenerable and schema-checked rather than an orphaned
   hand-edited YAML. Three bases (``site`` / ``upstream`` / ``source``) — see
   :class:`~watermark.greenops.model.WueBenchmark`; the derivation selects the AWS row
   because we are a tenant, not an operator (#1643/F4).

**Discipline.** Every figure in both tables is ``reference`` — an authoritative published
factor, *not* a metered fact about our own consumption. The derivation applies them;
PUE / utilization / which-subregion-a-workload-runs-in stay stated ``assumption``s
upstream. Nothing here is ``connector``-``verified``.
"""

from __future__ import annotations

import io
import os
import tempfile
from pathlib import Path
from typing import Any, cast

import httpx
import openpyxl
import yaml

from watermark.config import Settings, get_settings
from watermark.connectors import cached_get, to_float
from watermark.greenops.model import (
    EgridFactors,
    EgridResourceShare,
    EgridSubregion,
    InferenceEnergyTable,
    WueBenchmark,
    WueTable,
)
from watermark.hydrology.model import ProvenancedValue
from watermark.logging import get_logger

from . import GreenopsOfflineError
from ._readme import write_reference_yaml

log = get_logger(__name__)

# eGRID subregion field codes (row 2 of the SRL<yy> sheet), selected BY NAME. ``SRC2ERTA``
# is the annual CO2-equivalent total output emission rate (lb/MWh) — the electricity→CO2e
# factor; ``SRTRPR`` is the total-renewables generation share (a 0-1 fraction). The per-fuel
# resource-mix columns (also 0-1 fractions) drive the source-mix breakdown.
_CODE_SUBRGN = "SUBRGN"
_CODE_NAME = "SRNAME"
_CODE_CO2E_RATE = "SRC2ERTA"
_CODE_RENEWABLE = "SRTRPR"

# The generation resource-mix columns: field code -> (fuel key, display label). Our
# renewable classification (below) is a declared modeling choice, not read from eGRID; it
# is cross-checked against eGRID's own SRTRPR total in the connector test.
_MIX_COLUMNS: list[tuple[str, str, str]] = [
    ("SRCLPR", "coal", "Coal"),
    ("SROLPR", "oil", "Oil"),
    ("SRGSPR", "gas", "Natural gas"),
    ("SRNCPR", "nuclear", "Nuclear"),
    ("SRHYPR", "hydro", "Hydro"),
    ("SRBMPR", "biomass", "Biomass"),
    ("SRWIPR", "wind", "Wind"),
    ("SRSOPR", "solar", "Solar"),
    ("SRGTPR", "geothermal", "Geothermal"),
    ("SROFPR", "other_fossil", "Other fossil"),
    ("SROPPR", "other", "Other/unknown"),
]
_RENEWABLE_FUELS = frozenset({"hydro", "biomass", "wind", "solar", "geothermal"})


class EgridFormatError(RuntimeError):
    """The eGRID workbook was missing the expected subregion sheet or a required field code.

    Raised when the ``SRL<yy>`` sheet or a selected column (``SUBRGN`` / ``SRC2ERTA`` / …)
    is absent — i.e. the vintage's format drifted — so the miss is loud, not a silent
    empty table (mirrors ``Eia861Error``).
    """


def _subregion_sheet(year: int) -> str:
    """The subregion-level sheet name for a vintage: ``SRL23`` for eGRID2023."""
    return f"SRL{year % 100:02d}"


def _download_workbook(settings: Settings) -> bytes:
    """Return the eGRID workbook bytes, downloading to the greenops cache on first use."""
    cache = settings.greenops_cache_dir / "egrid" / f"egrid{settings.egrid_year}.xlsx"
    if cache.is_file():
        return cache.read_bytes()
    log.info("greenops.egrid.download", url=settings.egrid_data_url, year=settings.egrid_year)
    resp = httpx.get(
        settings.egrid_data_url,
        timeout=max(settings.greenops_request_timeout_s, 120.0),
        follow_redirects=True,
    )
    resp.raise_for_status()  # an EPA error page must not be cached as if it were the workbook
    data = resp.content
    cache.parent.mkdir(parents=True, exist_ok=True)
    # Unique sibling temp file (not a shared ``.part``) so concurrent downloads of the
    # same key don't clobber each other; atomically replaced into place, cleaned up on error.
    fd, tmp_name = tempfile.mkstemp(dir=cache.parent, prefix=f"{cache.name}.", suffix=".part")
    tmp = Path(tmp_name)
    try:
        with os.fdopen(fd, "wb") as out:
            out.write(data)
        os.replace(tmp, cache)
    finally:
        tmp.unlink(missing_ok=True)
    return data


def _header_index(ws: Any) -> tuple[int, dict[str, int]]:
    """Find the field-code header row: its 1-based row number + a field-code→column map.

    eGRID sheets carry a descriptive row 1 and the short field codes on row 2; scanning
    (rather than assuming row 2) tolerates a preamble-row shift between vintages.
    """
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
        codes = {str(c).strip(): i for i, c in enumerate(row) if c is not None}
        if _CODE_SUBRGN in codes:
            return row_num, codes
    raise EgridFormatError(f"no header row with {_CODE_SUBRGN!r} in the eGRID subregion sheet")


def _reduce_subregions(ws: Any, header: tuple[int, dict[str, int]]) -> list[dict[str, Any]]:
    """Reduce the subregion worksheet to per-subregion factor rows, selecting columns by name.

    ``header`` is ``(header_row_1based, field_code→column)`` from :func:`_header_index`; data
    rows are the ones below it carrying a subregion acronym. Missing required codes raise
    :class:`EgridFormatError`; a ``'--'`` (eGRID's N/A) coerces to 0.
    """
    header_row, index = header
    # Every selected code must be present, including each SR*PR mix column: a missing/renamed
    # field must be a loud EgridFormatError, not a silent 0.0 flattening the resource mix.
    required_codes = [
        _CODE_NAME,
        _CODE_CO2E_RATE,
        _CODE_RENEWABLE,
        *(c for c, _, _ in _MIX_COLUMNS),
    ]
    for required in required_codes:
        if required not in index:
            raise EgridFormatError(f"eGRID subregion sheet missing required field {required!r}")

    def cell(row: tuple[Any, ...], code: str) -> Any:
        col = index.get(code)
        return row[col] if col is not None and col < len(row) else None

    out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        code = cell(row, _CODE_SUBRGN)
        if not code or not str(code).strip():
            continue
        mix = [
            {
                "fuel": fuel,
                "label": label,
                "percent": round(to_float(cell(row, col), 0.0) * 100.0, 2),
                "renewable": fuel in _RENEWABLE_FUELS,
            }
            for col, fuel, label in _MIX_COLUMNS
        ]
        out.append(
            {
                "code": str(code).strip(),
                "name": str(cell(row, _CODE_NAME) or "").strip(),
                "co2e_lb_mwh": round(to_float(cell(row, _CODE_CO2E_RATE), 0.0), 3),
                "renewable_percent": round(to_float(cell(row, _CODE_RENEWABLE), 0.0) * 100.0, 2),
                "resource_mix": mix,
            }
        )
    if not out:
        raise EgridFormatError("eGRID subregion sheet had no data rows")
    return out


def _fetch_egrid(settings: Settings) -> dict[str, Any]:
    """Pull + reduce the eGRID subregion sheet (cached). The reduced rows are the payload."""
    year = settings.egrid_year
    key_params: dict[str, Any] = {
        "connector": "egrid",
        "report": "subregion_factors",
        "year": year,
        "sheet": _subregion_sheet(year),
    }

    def fetch() -> Any:
        wb = openpyxl.load_workbook(
            io.BytesIO(_download_workbook(settings)), read_only=True, data_only=True
        )
        sheet = _subregion_sheet(year)
        if sheet not in wb.sheetnames:
            wb.close()
            raise EgridFormatError(f"eGRID {year} workbook has no {sheet!r} sheet")
        ws = wb[sheet]
        rows = _reduce_subregions(ws, _header_index(ws))
        wb.close()
        return {"year": year, "vintage": f"eGRID{year}", "subregions": rows}

    return cast(
        "dict[str, Any]",
        cached_get(
            "egrid",
            key_params,
            fetch,
            cache_dir=settings.greenops_cache_dir,
            offline=settings.greenops_offline,
            fixtures_dir=settings.greenops_fixtures_dir,
            ttl_hours=settings.greenops_cache_ttl_hours,
            offline_error=GreenopsOfflineError,
        ),
    )


def build_egrid_factors(payload: dict[str, Any], *, source_url: str) -> EgridFactors:
    """Reduce the raw eGRID payload into an :class:`EgridFactors`. Pure over the payload.

    Every figure is ``reference`` (an authoritative published factor). Subregions are
    ranked by acronym; each subregion's resource mix keeps only fuels with a nonzero share,
    ranked by share descending.
    """
    vintage = str(payload.get("vintage") or f"eGRID{payload.get('year')}")
    cite = f"EPA {vintage} subregion file (SRL sheet)"
    subregions = [
        EgridSubregion(
            code=str(sr["code"]),
            name=str(sr["name"]),
            co2e_rate=ProvenancedValue.from_reference(
                float(sr["co2e_lb_mwh"]), "lb CO2e/MWh", cite, confidence="high"
            ),
            renewable_pct=ProvenancedValue.from_reference(
                float(sr["renewable_percent"]), "%", cite, confidence="high"
            ),
            resource_mix=sorted(
                (
                    EgridResourceShare(
                        fuel=str(m["fuel"]),
                        label=str(m["label"]),
                        percent=float(m["percent"]),
                        renewable=bool(m["renewable"]),
                    )
                    for m in sr["resource_mix"]
                    if float(m["percent"]) > 0
                ),
                key=lambda s: s.percent,
                reverse=True,
            ),
        )
        for sr in sorted(payload.get("subregions") or [], key=lambda s: str(s["code"]))
    ]
    return EgridFactors(
        year=int(payload["year"]),
        vintage=vintage,
        source_url=source_url,
        subregions=subregions,
        note=(
            f"EPA {vintage} subregion file: annual CO2-equivalent total output emission rate "
            "(SRC2ERTA, lb/MWh) + generation resource mix (SR*PR), read by field code from "
            "the workbook's own header row. Every figure is `reference` (an authoritative "
            "published factor), never metered. The renewable classification "
            "(hydro/biomass/wind/solar/geothermal) is our declared grouping, cross-checked "
            "against eGRID's own total-renewables share (SRTRPR)."
        ),
    )


def fetch_egrid_factors(*, settings: Settings | None = None) -> EgridFactors:
    """Pull + reduce the EPA eGRID subregion factor table (cached).

    Reads the vintage configured by ``settings.egrid_year`` / ``egrid_data_url``. An offline
    cache/fixture miss raises :class:`GreenopsOfflineError` naming the key to record.
    """
    settings = settings or get_settings()
    payload = _fetch_egrid(settings)
    return build_egrid_factors(payload, source_url=settings.egrid_data_url)


# --- the committed WUE benchmark table (EPRI / Uptime Institute) -----------------------------
# Hand-curated from published data-center water benchmarks (site WUE) plus the upstream
# water-for-electricity intensity (source WUE). These are representative published figures —
# `reference`, never a metered fact about our own cooling — transcribed with their source and
# vintage. The derivation (#1083) multiplies electricity by the site + source benchmarks for
# the modeled facility type; it stays a modeled `derived` number, never `verified`.

WUE_VINTAGE = "AWS 2025 / EPRI 2024 / Uptime Institute 2023 / Macknick et al. 2012"

# Every row carries a *dated* citation and a band (#1643/F5) — a benchmark quoted as a bare
# scalar hid a spread wide enough to change the answer. `basis` is three-way (#1643/F4):
# `site` (cooling per IT-kWh), `upstream` (the generation increment per facility-kWh — a term
# you ADD to a site WUE, not a WUE), `source` (an already-complete site+upstream figure).
_WUE_ROWS: list[dict[str, Any]] = [
    {
        "facility_type": "aws_published_site",
        "label": "AWS data centers, published site WUE",
        "value": 0.12,
        "low": 0.12,
        "high": 0.84,
        "basis": "site",
        "confidence": "medium",
        "cite": "AWS published global data-center WUE — 0.12 L/kWh water withdrawn per kWh "
        "of IT load (2025; 0.15 in 2024), sustainability.aboutamazon.com/products-services/"
        "aws-cloud, accessed 2026-07-31. Band: AWS's global average (low) to the "
        "industry-average withdrawal WUE of 0.84 L/kWh it compares itself against (high).",
        "note": "The row the derivation selects, because our compute is an AWS tenancy: "
        "the cooling water is drawn by AWS's facility, apportioned to us by billed IT-kWh. "
        "A global average — a specific region's WUE can be several times either bound, "
        "which is what the band carries.",
    },
    {
        "facility_type": "industry_average_site",
        "label": "Data center, industry-average site WUE",
        "value": 1.8,
        "low": 0.84,
        "high": 2.2,
        "basis": "site",
        "confidence": "low",
        "cite": "Commonly-quoted industry-average data-center site WUE (~1.8 L/kWh); no "
        "single dated primary source — the figure recurs across Uptime Institute survey "
        "commentary and secondary literature. Band: AWS's stated industry-average "
        "withdrawal WUE 0.84 L/kWh (2025) to the EPRI evaporative row below.",
        "note": "Kept for comparison, NOT applied: it is an undated composite whose "
        "boundary (withdrawal vs consumption, cooling-only vs total) is not stated in any "
        "one primary source, hence low confidence and a band wider than its own value.",
    },
    {
        "facility_type": "hyperscale_evaporative",
        "label": "Hyperscale, evaporative-tower cooling",
        "value": 2.2,
        "low": 1.8,
        "high": 3.0,
        "basis": "site",
        "confidence": "low",
        "cite": "EPRI data-center water-use characterization — evaporative (open-loop "
        "cooling-tower) facilities; a representative published upper range, not a dated "
        "single measurement.",
        "note": "Wet cooling trades electricity for water; a representative upper-range "
        "site WUE for evaporatively cooled campuses.",
    },
    {
        "facility_type": "closed_loop_airside",
        "label": "Closed-loop / air-side economized cooling",
        "value": 0.3,
        "low": 0.05,
        "high": 0.6,
        "basis": "site",
        "confidence": "low",
        "cite": "EPRI / industry closed-loop + air-side-economizer data-center WUE; a "
        "representative published range, not a dated single measurement.",
        "note": "Dry / closed-loop cooling draws little on-site water but more electricity; "
        "a representative low-water site WUE.",
    },
    {
        "facility_type": "grid_upstream",
        "label": "Grid generation, upstream increment (water for electricity)",
        "value": 1.9,
        "low": 0.78,
        "high": 2.54,
        "basis": "upstream",
        "confidence": "low",
        "cite": "Macknick, Newmark, Heath & Hallett, 'Operational water consumption and "
        "withdrawal factors for electricity generating technologies', Environ. Res. Lett. 7 "
        "045802 (2012-12-20), median operational water-CONSUMPTION factors. Central: US "
        "thermoelectric fleet average. Band: gas combined-cycle with cooling tower "
        "205 gal/MWh = 0.78 L/kWh (low) to nuclear with cooling tower 672 gal/MWh = "
        "2.54 L/kWh (high).",
        "note": "Water consumed upstream to generate 1 kWh DELIVERED — the increment added "
        "to a site WUE to reach a source WUE, never a WUE on its own. It applies to the "
        "whole facility draw (IT x PUE), not the IT load. The band is the generation-mix "
        "spread the subregion (eGRID) determines.",
    },
]


def build_wue_table() -> WueTable:
    """The committed WUE benchmark table, assembled from the in-code canonical (all ``reference``)."""
    return WueTable(
        vintage=WUE_VINTAGE,
        benchmarks=[
            WueBenchmark(
                facility_type=str(r["facility_type"]),
                label=str(r["label"]),
                wue=ProvenancedValue.from_reference(
                    float(r["value"]),
                    "L/kWh",
                    str(r["cite"]),
                    confidence=r["confidence"],
                    low=float(r["low"]),
                    high=float(r["high"]),
                ),
                basis=str(r["basis"]),
                note=str(r["note"]),
            )
            for r in _WUE_ROWS
        ],
        note=(
            "Water Usage Effectiveness benchmarks (liters of water per kWh), hand-curated "
            f"from published sources ({WUE_VINTAGE}); every row carries a dated citation and "
            "a band. THREE bases, and the distinction is load-bearing: SITE is facility "
            "cooling water per kWh of IT load; UPSTREAM is the increment consumed generating "
            "the electricity delivered to the whole facility (per facility-kWh) — a term you "
            "ADD to a site WUE, not a WUE on its own; SOURCE is an already-complete "
            "site+upstream figure. Site + upstream compose; site + source double-counts "
            "cooling and must never be summed. Every figure is `reference` (a published "
            "benchmark), never a metered fact about our own cooling; the derivation (#1083) "
            "applies them and stays `derived`. Representative rows are marked low-confidence."
        ),
    )


# --- committed reference artifacts (data/reference/greenops/factors/*.yaml) ------------------

_FACTORS_RELDIR = Path("reference") / "greenops" / "factors"


def _egrid_relpath(year: int) -> Path:
    return _FACTORS_RELDIR / f"egrid-{year}.yaml"


_WUE_RELPATH = _FACTORS_RELDIR / "wue-benchmarks.yaml"

_FACTORS_README = """\
# GreenOps — carbon-intensity & water factor tables

The `reference` factor tables the footprint derivation (#1083) multiplies our electricity
figure by, so the electricity → CO2e / source-mix / water conversion reads authoritative
published factors rather than a hand-maintained constant. Every figure here is
`source: reference` — an authoritative published factor, **not** a metered fact about our
own consumption. Nothing here is `verified`; the derivation that applies these stays
`derived`, and PUE / utilization / which-subregion-a-workload-runs-in remain stated
`assumption`s upstream.

## EPA eGRID subregion factors (`egrid-<year>.yaml`, #1082)

- **Source:** EPA Emissions & Generation Resource Integrated Database (eGRID), the annual
  subregion file (`SRL<yy>` sheet). Regenerate with `watermark greenops egrid --write`
  (public workbook, no API key). The raw ~20 MB xlsx caches under
  `data/cache/greenops/egrid/` (git-ignored); only the reduced subregion rows are
  committed.
- **Vintage:** set by `WATERMARK_EGRID_YEAR` / `egrid_data_url` (default eGRID2023). A new
  vintage is a config change: bump the year and point the URL at that release's workbook.
- **Fields** (selected by field code from the sheet's own header row, never by index):
  `SRC2ERTA` — annual CO2-equivalent total output emission rate (lb/MWh, the
  electricity→CO2e factor); `SRTRPR` — total-renewables generation share; the `SR*PR`
  columns — the per-fuel generation mix. The renewable grouping
  (hydro/biomass/wind/solar/geothermal) is our declared classification, cross-checked
  against eGRID's own `SRTRPR` total.

## WUE benchmarks (`wue-benchmarks.yaml`, #1082)

- **Source:** hand-curated Water Usage Effectiveness benchmarks (liters of water per kWh)
  from published data-center water studies (AWS's own published WUE, EPRI, Uptime Institute)
  plus the upstream water-for-electricity intensity (Macknick et al. 2012). Not a live pull —
  an in-code canonical emitted by `watermark greenops egrid --write`, so it stays regenerable
  and schema-checked. Every row carries a dated citation and a band (#1643/F5).
- **Bases — three, and the distinction is load-bearing:** `site` is facility cooling water
  per kWh of **IT** load; `upstream` is the increment consumed generating the electricity
  delivered to the **whole facility** (IT x PUE) — a term you *add* to a site WUE, not a WUE
  on its own; `source` is an already-complete site+upstream figure. Site + upstream compose
  into a source-basis total; site + source double-counts cooling and must never be summed.
- **Which row applies:** the derivation selects `aws_published_site`, because the platform
  runs no data center — the cooling water is AWS's, apportioned to us by billed IT-kWh
  (#1643/F4). `industry_average_site` is kept for comparison and deliberately not applied.

## Inference energy (`inference-energy.yaml`, #1643/F2)

- **Source:** hand-curated per-1,000-**output**-token energy coefficients by model class,
  from published third-party measurements (Epoch AI 2025-02, Jegham et al. 2025-05, Google
  2025-08). No provider publishes a per-token energy figure for its hosted models, so this
  is a `reference` table of outside estimates, never a metered fact. Emitted alongside the
  other factor tables by `watermark greenops egrid --write`.
- **Banded on purpose:** published estimates of the same quantity span roughly an order of
  magnitude, driven mostly by boundary — an accelerator-only figure excludes host CPU/DRAM,
  idle capacity and facility overhead, and Google's 2025-08 production measurement found the
  accelerator is only 58% of full-stack energy.
- **Basis:** output tokens. Decode dominates inference energy; prefill is far cheaper per
  token, so applying these to an input+output total would overstate a cache-heavy agentic
  workload several-fold. An unmapped model id is priced at `default_class`, never dropped.
"""


def _write_yaml(out: Path, table: EgridFactors | WueTable | InferenceEnergyTable) -> Path:
    # The factor tables live in their own `factors/` dir, so they carry _FACTORS_README rather
    # than the shared `data/reference/greenops/` one.
    return write_reference_yaml(
        out, table, log=log, event="greenops.egrid.wrote", readme=_FACTORS_README
    )


def write_egrid_factors(factors: EgridFactors, *, settings: Settings | None = None) -> Path:
    """Persist the eGRID factor table as committed reference YAML + the factors README."""
    settings = settings or get_settings()
    return _write_yaml(settings.data_dir / _egrid_relpath(factors.year), factors)


def write_wue_table(table: WueTable, *, settings: Settings | None = None) -> Path:
    """Persist the WUE benchmark table as committed reference YAML + the factors README."""
    settings = settings or get_settings()
    return _write_yaml(settings.data_dir / _WUE_RELPATH, table)


def write_inference_energy(
    table: InferenceEnergyTable, *, settings: Settings | None = None
) -> Path:
    """Persist the inference-energy table as committed reference YAML + the factors README.

    Lives here rather than beside its builder so all three factor tables share one writer and
    one README (the folder's contract is written once, not three times).
    """
    from .inference_energy import inference_energy_path

    return _write_yaml(inference_energy_path(settings), table)


def load_egrid_factors(path: Path) -> EgridFactors:
    """Load a committed :class:`EgridFactors` YAML."""
    return EgridFactors.model_validate(yaml.safe_load(path.read_text(encoding="utf-8")))


def load_wue_table(path: Path) -> WueTable:
    """Load a committed :class:`WueTable` YAML."""
    return WueTable.model_validate(yaml.safe_load(path.read_text(encoding="utf-8")))
