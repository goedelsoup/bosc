"""EPA eGRID connector + WUE factor table (#1082): the subregion reduce selects columns by
field code (not index), the renewable classification cross-checks eGRID's own total, every
figure is `reference` (never metered), an offline miss / format drift fail loudly, and the
committed factor artifacts validate + round-trip.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from watermark.config import Settings
from watermark.greenops.connectors import (
    EgridFormatError,
    GreenopsOfflineError,
    build_egrid_factors,
    build_inference_energy_table,
    build_wue_table,
    fetch_egrid_factors,
    load_egrid_factors,
    load_inference_energy,
    load_wue_table,
    write_egrid_factors,
    write_wue_table,
)
from watermark.greenops.connectors.egrid import (
    _header_index,
    _reduce_subregions,
    _subregion_sheet,
)

REPO_ROOT = Path(__file__).resolve().parents[1]
FACTORS = REPO_ROOT / "data" / "reference" / "greenops" / "factors"

# The renewable-fuel keys the connector classifies as renewable (must match egrid._RENEWABLE_FUELS).
_RENEWABLE = {"hydro", "biomass", "wind", "solar", "geothermal"}


def test_egrid_reduce_is_faithful(greenops_settings: Settings) -> None:
    factors = fetch_egrid_factors(settings=greenops_settings)
    assert factors.vintage == "eGRID2023"
    assert factors.year == 2023
    # All 27 published subregions, ranked by acronym.
    assert len(factors.subregions) == 27
    assert [s.code for s in factors.subregions] == sorted(s.code for s in factors.subregions)

    by_code = factors.by_code()
    # RFCW (RFC West) is Lima's LOCAL grid — it is NOT where our own compute runs, and the
    # footprint derivation never reads it (that is SRVC, us-east-1; #1643/F5). Asserted here
    # only because it is a stable, high-emission row of the real workbook.
    rfcw = by_code["RFCW"]
    assert rfcw.name == "RFC West"
    assert rfcw.co2e_rate.value == pytest.approx(916.054)
    assert rfcw.co2e_rate.unit == "lb CO2e/MWh"
    assert rfcw.renewable_pct.value == pytest.approx(7.6)


def test_resource_mix_ranked_and_renewables_cross_check(greenops_settings: Settings) -> None:
    factors = fetch_egrid_factors(settings=greenops_settings)
    for sr in factors.subregions:
        # Mix keeps only nonzero shares, ranked descending.
        percents = [m.percent for m in sr.resource_mix]
        assert percents == sorted(percents, reverse=True)
        assert all(m.percent > 0 for m in sr.resource_mix)
        # Our renewable grouping matches eGRID's own total-renewables share (SRTRPR).
        renewable_from_mix = sum(m.percent for m in sr.resource_mix if m.renewable)
        assert renewable_from_mix == pytest.approx(sr.renewable_pct.value, abs=0.5)
        assert all((m.fuel in _RENEWABLE) == m.renewable for m in sr.resource_mix)


def test_every_egrid_figure_is_reference_never_metered(greenops_settings: Settings) -> None:
    factors = fetch_egrid_factors(settings=greenops_settings)
    values = factors.all_values()
    assert values
    assert all(v.source == "reference" for v in values)
    assert not any(v.verified for v in values)


def _synthetic_subregion_sheet() -> object:
    """A tiny eGRID-shaped sheet: a descriptive row 1, field codes on row 2, one data row.

    Column order is deliberately shuffled so the test proves selection is BY NAME, not index.
    """
    # Field codes deliberately shuffled + the full SR*PR mix set (all required up front). The
    # named identity/rate/renewable columns sit next to unrelated shares to prove by-name reads.
    codes = ["SRNAME", "SRGSPR", "SUBRGN", "SRC2ERTA", "SRTRPR", "SRWIPR"]
    data: list[object] = ["Test Region", 0.60, "TEST", 500.0, 0.40, 0.40]
    for code in (
        "SRCLPR",
        "SROLPR",
        "SRNCPR",
        "SRHYPR",
        "SRBMPR",
        "SRSOPR",
        "SRGTPR",
        "SROFPR",
        "SROPPR",
    ):
        codes.append(code)
        data.append(0.0)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["desc"] * len(codes))  # row 1 (ignored)
    ws.append(codes)  # row 2: field codes
    ws.append(data)  # data
    ws.append([None] * len(codes))  # trailing blank
    return ws


def test_reduce_selects_columns_by_name_not_index() -> None:
    ws = _synthetic_subregion_sheet()
    index = _header_index(ws)
    rows = _reduce_subregions(ws, index)
    assert len(rows) == 1
    row = rows[0]
    assert row["code"] == "TEST"
    assert row["name"] == "Test Region"
    assert row["co2e_lb_mwh"] == pytest.approx(500.0)  # not the 0.60 gas share next to it
    assert row["renewable_percent"] == pytest.approx(40.0)  # SRTRPR 0.40 -> 40%
    mix = {m["fuel"]: m["percent"] for m in row["resource_mix"]}
    assert mix["gas"] == pytest.approx(60.0)
    assert mix["wind"] == pytest.approx(40.0)


def test_reduce_raises_on_missing_field_code() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["SUBRGN", "SRNAME"])  # no SRC2ERTA / SRTRPR
    ws.append(["TEST", "Test Region"])
    with pytest.raises(EgridFormatError):
        _reduce_subregions(ws, _header_index(ws))


def test_reduce_raises_on_missing_mix_column() -> None:
    # A renamed/absent SR*PR mix column must fail loud, not silently flatten the mix to 0.
    ws = _synthetic_subregion_sheet()
    header_row, index = _header_index(ws)
    del index["SRCLPR"]  # simulate coal-share column drift
    with pytest.raises(EgridFormatError, match="SRCLPR"):
        _reduce_subregions(ws, (header_row, index))


def test_header_index_raises_without_subrgn() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["not", "an", "egrid", "sheet"])
    with pytest.raises(EgridFormatError):
        _header_index(ws)


def test_offline_miss_raises(greenops_settings: Settings) -> None:
    # A vintage with no committed fixture -> an actionable offline miss.
    settings = greenops_settings.model_copy(update={"egrid_year": 1999})
    with pytest.raises(GreenopsOfflineError):
        fetch_egrid_factors(settings=settings)


def test_subregion_sheet_name() -> None:
    assert _subregion_sheet(2023) == "SRL23"
    assert _subregion_sheet(2020) == "SRL20"


# --- WUE benchmark table ----------------------------------------------------------------------


def test_wue_table_is_all_reference() -> None:
    table = build_wue_table()
    assert table.benchmarks
    values = table.all_values()
    assert all(v.source == "reference" for v in values)
    assert not any(v.verified for v in values)
    assert all(v.unit == "L/kWh" for v in values)
    # Every row carries a band and a dated citation (#1643/F5) — a bare scalar hid a spread
    # wide enough to change the answer.
    assert all(v.low is not None and v.high is not None for v in values)
    assert all(v.low_or_value <= v.value <= v.high_or_value for v in values)

    # `site` and `upstream` are the two bases the derivation composes (#1643/F4). No row is
    # typed `source`: an already-complete site+upstream figure added to a site row would
    # double-count cooling, which is the sum the model's own docstring forbade.
    assert {b.basis for b in table.benchmarks} == {"site", "upstream"}
    upstream = [b for b in table.benchmarks if b.basis == "upstream"]
    assert [b.facility_type for b in upstream] == ["grid_upstream"]

    # The row the derivation actually selects is our PROVIDER's published figure, not an
    # industry average standing in for a data center we do not operate.
    aws = next(b for b in table.benchmarks if b.facility_type == "aws_published_site")
    assert aws.basis == "site"
    assert "accessed 2026-07-31" in (aws.wue.citation or "")


def test_inference_energy_table_is_banded_reference_priced_on_output_tokens() -> None:
    """#1643/F2: the coefficients that put inference in the energy chain.

    No provider publishes a per-token energy figure for its hosted models, so every row is a
    `reference` transcription of a third-party measurement — banded, because the published
    estimates of the same quantity span roughly an order of magnitude, and priced on OUTPUT
    tokens, because decode dominates and applying it to an input+output total would overstate a
    cache-heavy agentic workload several-fold.
    """
    table = build_inference_energy_table()
    values = table.all_values()
    assert values
    assert all(v.source == "reference" for v in values)
    assert not any(v.verified for v in values)
    assert all(v.unit == "Wh/1k output tokens" for v in values)
    assert all(v.low is not None and v.high is not None for v in values)
    assert all(b.basis == "output_tokens" for b in table.benchmarks)

    # Model ids resolve by longest prefix, so a dated point release needs no new entry...
    assert table.benchmark_for("claude-opus-4-8-20260115").model_class == "frontier"
    assert table.benchmark_for("claude-sonnet-4-6").model_class == "mid_tier"
    assert table.benchmark_for("claude-haiku-4-5").model_class == "small"
    # ...and an unmapped id is priced at the conservative default rather than dropped.
    assert table.benchmark_for("some-unknown-model").model_class == table.default_class
    by_class = table.by_class()
    assert by_class["frontier"].wh_per_1k_tokens.value > by_class["mid_tier"].wh_per_1k_tokens.value


def test_committed_inference_energy_artifact_matches_canonical() -> None:
    assert (
        load_inference_energy(FACTORS / "inference-energy.yaml") == build_inference_energy_table()
    )


def test_wue_table_round_trips(tmp_path: Path) -> None:
    table = build_wue_table()
    settings = Settings(data_dir=tmp_path)
    out = write_wue_table(table, settings=settings)
    assert load_wue_table(out) == table


def test_egrid_factors_round_trip(greenops_settings: Settings, tmp_path: Path) -> None:
    factors = fetch_egrid_factors(settings=greenops_settings)
    settings = Settings(data_dir=tmp_path)
    out = write_egrid_factors(factors, settings=settings)
    assert load_egrid_factors(out) == factors


# --- the committed factor artifacts -----------------------------------------------------------


def test_committed_egrid_artifact_validates() -> None:
    factors = load_egrid_factors(FACTORS / "egrid-2023.yaml")
    assert factors.vintage == "eGRID2023"
    assert len(factors.subregions) == 27
    assert all(v.source == "reference" for v in factors.all_values())


def test_committed_wue_artifact_matches_canonical() -> None:
    assert load_wue_table(FACTORS / "wue-benchmarks.yaml") == build_wue_table()


def test_build_egrid_factors_carries_source_url() -> None:
    payload = {
        "year": 2023,
        "vintage": "eGRID2023",
        "subregions": [
            {
                "code": "TEST",
                "name": "Test Region",
                "co2e_lb_mwh": 500.0,
                "renewable_percent": 40.0,
                "resource_mix": [
                    {"fuel": "gas", "label": "Natural gas", "percent": 60.0, "renewable": False},
                    {"fuel": "wind", "label": "Wind", "percent": 40.0, "renewable": True},
                ],
            }
        ],
    }
    factors = build_egrid_factors(payload, source_url="https://example.test/egrid.xlsx")
    assert factors.source_url == "https://example.test/egrid.xlsx"
    assert factors.by_code()["TEST"].co2e_rate.value == pytest.approx(500.0)
